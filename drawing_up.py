from openpyxl import Workbook
from slots_sort import sort_slots
import re

def alg1(stafftmp, staffdatatmp, slotstmp, sh1tmp):
    smeni = []
    for i in range(0,staff):
        smeni.append(0)
    for i in range(0, slotstmp):
        smeni1 = []
        for j in range(0, staff):
            smeni1.append(0)
        for j in range(0, stafftmp):
            if int(staffdatatmp[j][i]) == 2:
                smeni1[j] += 1 + smeni[j]
            if smeni1[j] == 0:
                smeni1[j] = 999

        smeni1_num = list (enumerate(smeni1, 0))

        r = min(smeni1_num, key=lambda g: g[1])

        del(smeni1_num[r[0]])
        if r[1]!=999:

            smeni[r[0]]+=1
        r2 = min(smeni1_num, key=lambda g: g[1])
        if r2[1] != 999 and r2[0]!=r[0]:
            smeni[r2[0]] += 1
            l = [i, r[0],r2[0]]

        else:
            l = [i,r[0],'-']
        if r[1]==999:
            l = [i,'-','-']
        sh1tmp.append(l)

    #sh1tmp.append(smeni)

def alg2(stafftmp, staffdatatmp, slotstmp, sh1tmp):
    smeni = []
    for i in range(0, staff):
        smeni.append(0)
    for i in range(0, slotstmp):
        smeni1 = []
        for j in range(0, staff):
            smeni1.append(0)

        for j in range(0, stafftmp):
            if int(staffdatatmp[j][i]) == 2 or int(staffdatatmp[j][i]) == 1:
                smeni1[j] += 1 + smeni[j]
            if smeni1[j] == 0:
                smeni1[j] = 999
        smeni1_num = list (enumerate(smeni1, 0))

        r = min(smeni1_num, key=lambda g: g[1])
        del(smeni1_num[r[0]])
        if r[1]!=999:
            smeni[r[0]]+=1
        r2 = min(smeni1_num, key=lambda g: g[1])
        if r2[1] != 999 and r2[0]!=r[0]:
            smeni[r2[0]] += 1
            l = [i, r[0],r2[0]]
        else:
            l = [i, r[0],'-']
        if r[1] == 999:
            l = [i, '-', '-']
        sh1tmp.append(l)

    #sh1tmp.append(smeni)

def func(thread_name, staff_priority, staff_flows):
    result = 0
    for staff in staff_priority:
        if thread_name in staff_flows[staff]:
            result += 1
    return result

if __name__ == '__main__':
    staff_schedule, staff_load, staff_threads, thread_names, staff_flows = [], [], [], [], []
    staff_number, threads_number = map(int, input().split())
    thread_names = re.sub(r'\s*','',input()).split(',')
    print(thread_names)
    for num in range(1, staff_number + 1):
        print(f"Enter {num} staff schedule and max: ", end='')
        line = input()
        line = re.sub(r'\s*,\s*', ',', line)
        staff_line, staff_max_load, staff_max_threads, staff_flows_str = line.split()
        staff_schedule.append(list(map(int, list(staff_line))))
        staff_load.append(int(staff_max_load))
        staff_threads.append(int(staff_max_threads))
        staff_flows.append(staff_flows_str.split(','))
        assert len(staff_schedule[-1]) == 28
    # print(staff_schedule)
    print(staff_flows)
    # print(staff_number)
    # print(threads_number)
    workbook = Workbook()
    worksheet = workbook.active

    for i in range(28):
        worksheet.cell(row=i+2, column=1).value = f"Смена {i+1}"
    for i in range(threads_number):
        worksheet.cell(row=1, column=i+2).value = thread_names[i]

    list_slots = sort_slots(staff_schedule)
    staff_sum = [0 for _ in range(staff_number)]
    staff_slots = [[] for _ in range(staff_number)]
    threads_sum = [0 for _ in range(threads_number)]
    for slot in list_slots:
        staff_1 = []
        staff_2 = []
        for staff in range(staff_number):
            if staff_schedule[staff][slot] == 1:
                staff_1.append(staff)
            if staff_schedule[staff][slot] == 2:
                staff_2.append(staff)
        staff_1 = sorted(staff_1, key=lambda staff: staff_sum[staff])
        staff_2 = sorted(staff_2, key=lambda staff: staff_sum[staff])
        for staff in staff_1.copy():
            for _ in range(staff_threads[staff] - 1):
                staff_1.append(staff)
        for staff in staff_2.copy():
            for _ in range(staff_threads[staff] - 1):
                staff_2.append(staff)
        staff_priority = staff_2 + staff_1
        staff_priority = list(filter(lambda staff: staff_sum[staff] < staff_load[staff], staff_priority))
        #print(staff_priority)
        staff_priority = list(
            filter(lambda staff: not ((slot - 1) in staff_slots[staff] and (slot - 2) in staff_slots[staff]),
                staff_priority))
        staff_priority = list(
            filter(lambda staff: not ((slot + 1) in staff_slots[staff] and (slot + 2) in staff_slots[staff]),
                 staff_priority))
        staff_priority = list(
            filter(lambda staff: not ((slot - 1) in staff_slots[staff] and (slot + 1) in staff_slots[staff]),
                staff_priority))
        if slot % 4 == 0:
            staff_priority = list(
                filter(lambda staff: not ((slot + 1) in staff_slots[staff]),
                        staff_priority))
        if slot % 4 == 1:
            staff_priority = list(
                filter(lambda staff: not ((slot - 1) in staff_slots[staff]),
                       staff_priority))
        #print(staff_priority)
        threads_priority = list(range(threads_number))
        threads_priority = sorted(threads_priority, key=lambda thread: threads_sum[thread])
        threads_priority = sorted(threads_priority, key=lambda thread: func(thread_names[thread], set(staff_priority), staff_flows))
        for thread in threads_priority:
            staff = None
            for i, staff_x in enumerate(staff_priority):
                if thread_names[thread] in staff_flows[staff_x]:
                    staff = staff_x
                    staff_priority.pop(i)
                    break
            if staff is not None:
                worksheet.cell(row=slot+2, column=thread+2).value = f" Сотрудник {staff+1}"
                staff_sum[staff] += 1
                staff_slots[staff].append(slot)
                threads_sum[thread] += 1
        print(staff_sum)

    workbook.save("result.xlsx")


# wb = Workbook()
#
# wb["Sheet"].title = "Schedule"
# wb2 = Workbook()
# wb2["Sheet"].title = "Schedule2"
# sh1 = wb.active
# sh2 = wb2.active
# print("Enter the number of staff")
# staff = int(input()) #количество человек
# threads = 6 #int(input()) #количество потоков(пока не использовал)
# slots = 28 #слоты
# staffdata = [] #список для каджого человека
# firstline = [staff, threads]
#
# for i in range(0,staff):
#     print("Enter ",i+1," staff")
#     tmp = []
#     tmp = input()
#     tmp2= []
#     if len(tmp)!=28:
#         print("You should enter 28 slots for ",i+1," staff")
#         break
#     for j in range(0,slots):
#         tmp2.append(int(tmp[j]))
#     staffdata.append(tmp2)
#
#
# alg1(staff,staffdata,slots,sh1)
# alg2(staff,staffdata,slots,sh2)
# wb.save("Schedule.xlsx")
# wb2.save("Schedule2.xlsx")
